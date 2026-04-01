import pandas as pd
import logging
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def write_dataframe_as_table(ws, df, sheet_title, start_row=1):
    """Escribe un DataFrame como tabla con encabezado formateado."""
    ws.title = sheet_title
    
    # Encabezado con fondo verde (#c6efce), texto oscuro, negrita
    header_fill = PatternFill("solid", fgColor="C6EFCE")
    header_font = Font(name="Arial", bold=True, color="000000", size=10)
    
    # Escribir encabezados
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir datos
    data_font = Font(name="Arial", size=10)
    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            # Alternar color de fila: blanco y verde muy claro
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="E2EFDA")
    
    # Ajustar ancho de columnas automáticamente
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        # Handle nan/empty carefully for length check
        max_len = max(len(str(col_name)), df[col_name].astype(str).map(len).max() if not df.empty else 0)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
        
    logger.info(f"Hoja '{sheet_title}' creada con {len(df)} filas escritas.")

def generate_excel_report(presentacion: pd.DataFrame, codigos_df: pd.DataFrame, 
                          azure_df: pd.DataFrame, tracking_df: pd.DataFrame,
                          output_path: str, start_date: str = "", end_date: str = ""):
    logger.info("Generando archivo XLSX...")
    
    wb = Workbook()
    
    # Hoja 1: HORAS PROYECTO
    ws_presentacion = wb.active
    
    if start_date and end_date:
        title = f"REPORTE DE HORAS ASIGNADAS A PROYECTOS PERIODO : {start_date} - {end_date}"
        ws_presentacion.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(presentacion.columns), 5))
        title_cell = ws_presentacion.cell(row=1, column=1, value=title)
        title_cell.font = Font(name="Arial", bold=True, size=11)
        start_row = 3
    else:
        start_row = 1

    write_dataframe_as_table(ws_presentacion, presentacion, "HORAS PROYECTO", start_row=start_row)
    
    # Formato especial para % ASIGNADO A PROYECTOS (la columna Suma ha sido deshabilitada del reporte)
    idx_pct = presentacion.columns.get_loc("% ASIGNADO A PROYECTOS") + 1
    # idx_suma = presentacion.columns.get_loc("Suma") + 1
    
    for row in ws_presentacion.iter_rows(min_row=start_row + 1, max_row=ws_presentacion.max_row):
        for cell in row:
            if cell.column == idx_pct:
                cell.number_format = "0%"
            # if cell.column == idx_suma:
            #     cell.number_format = "0.0"

    # Hoja 2: CODIGOS PROYECTOS
    ws_codigos = wb.create_sheet()
    write_dataframe_as_table(ws_codigos, codigos_df, "CODIGOS PROYECTOS")
    
    # Hoja 3: DETALLE - AZURE
    ws_azure = wb.create_sheet()
    write_dataframe_as_table(ws_azure, azure_df, "DETALLE - AZURE")
    
    # [COMENTADO TEMPORALMENTE: INTEGRACION TRACKINGTIME]
    # Para volver a integrar TrackingTime, descomentar el siguiente bloque:
    # Hoja 4: DETALLE - TRACKINGTIME
    # ws_tracking = wb.create_sheet()
    # write_dataframe_as_table(ws_tracking, tracking_df, "DETALLE - TRACKINGTIME")
    
    # Guardar archivo
    wb.save(output_path)
    logger.info(f"Archivo Excel generado exitosamente en: {output_path}")

